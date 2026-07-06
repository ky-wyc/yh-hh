from __future__ import annotations

import csv
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from pathlib import PurePath
from typing import Any

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".txt", ".md", ".csv", ".xlsx", ".xlsm"}
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_DOCUMENT_CHARS = 20_000
ROWS_PER_DOCUMENT = 100


@dataclass(slots=True)
class ImportedKnowledgeDocument:
    title: str
    content: str


@dataclass(slots=True)
class KnowledgeImportReport:
    file_name: str
    file_type: str
    source_count: int = 0
    row_count: int = 0
    imported_row_count: int = 0
    skipped_empty_rows: int = 0
    document_count: int = 0
    truncated: bool = False
    warnings: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ImportedKnowledge:
    title: str
    content: str
    file_type: str
    documents: list[ImportedKnowledgeDocument]
    report: KnowledgeImportReport


def parse_imported_knowledge(filename: str, data: bytes, title: str = "") -> ImportedKnowledge:
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("file is larger than 5MB")
    suffix = PurePath(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError("unsupported knowledge file type")

    base_title = (title.strip() or PurePath(filename or "imported-document").stem)[:255]
    base_title = base_title or "imported-document"
    report = KnowledgeImportReport(
        file_name=filename or "",
        file_type=suffix.lstrip("."),
    )

    if suffix in {".txt", ".md"}:
        documents = parse_text_documents(data, base_title, report)
    elif suffix == ".csv":
        documents = parse_csv_documents(data, base_title, report)
    else:
        documents = parse_workbook_documents(data, base_title, report)

    documents = [document for document in documents if document.content.strip()]
    if not documents:
        raise ValueError("imported file has no readable text")
    report.document_count = len(documents)
    return ImportedKnowledge(
        title=documents[0].title,
        content=documents[0].content,
        file_type=report.file_type,
        documents=documents,
        report=report,
    )


def parse_text_documents(
    data: bytes,
    base_title: str,
    report: KnowledgeImportReport,
) -> list[ImportedKnowledgeDocument]:
    content = normalize_content(decode_text(data))
    if not content:
        return []
    report.source_count = 1
    sections = split_text(content, MAX_DOCUMENT_CHARS)
    return [
        ImportedKnowledgeDocument(
            title=part_title(base_title, index, len(sections)),
            content=section,
        )
        for index, section in enumerate(sections, start=1)
    ]


def parse_csv_documents(
    data: bytes,
    base_title: str,
    report: KnowledgeImportReport,
) -> list[ImportedKnowledgeDocument]:
    reader = csv.reader(StringIO(decode_text(data)))
    header: list[str] = []
    row_sections: list[tuple[int, str]] = []
    for row_number, row in enumerate(reader, start=1):
        values = [str(cell).strip() for cell in row]
        if not any(values):
            report.skipped_empty_rows += 1
            continue
        report.row_count += 1
        if not header:
            header = values
            continue
        row_sections.append((row_number, format_structured_row(row_number, values, header)))
        report.imported_row_count += 1
    if not row_sections and header:
        row_sections.append((1, "Row 1\n" + "\n".join(value for value in header if value)))
        report.imported_row_count = 1
    report.source_count = 1
    return pack_row_sections(
        base_title=base_title,
        source_label="CSV",
        row_sections=row_sections,
    )


def parse_workbook_documents(
    data: bytes,
    base_title: str,
    report: KnowledgeImportReport,
) -> list[ImportedKnowledgeDocument]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    documents: list[ImportedKnowledgeDocument] = []
    try:
        report.source_count = len(workbook.worksheets)
        for sheet in workbook.worksheets:
            header: list[str] = []
            row_sections: list[tuple[int, str]] = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [format_cell(value) for value in row]
                if not any(values):
                    report.skipped_empty_rows += 1
                    continue
                report.row_count += 1
                if not header:
                    header = values
                    continue
                row_sections.append((row_number, format_structured_row(row_number, values, header)))
                report.imported_row_count += 1
            if not row_sections and header:
                row_sections.append((1, "Row 1\n" + "\n".join(value for value in header if value)))
                report.imported_row_count += 1
            documents.extend(
                pack_row_sections(
                    base_title=base_title,
                    source_label=f"Sheet {sheet.title}",
                    row_sections=row_sections,
                )
            )
    finally:
        workbook.close()
    return documents


def pack_row_sections(
    *,
    base_title: str,
    source_label: str,
    row_sections: list[tuple[int, str]],
) -> list[ImportedKnowledgeDocument]:
    documents: list[ImportedKnowledgeDocument] = []
    current: list[str] = []
    start_row = 0
    current_chars = 0
    current_count = 0
    for row_number, section in row_sections:
        section_chars = len(section) + 2
        should_flush = bool(current) and (
            current_count >= ROWS_PER_DOCUMENT or current_chars + section_chars > MAX_DOCUMENT_CHARS
        )
        if should_flush:
            documents.append(make_row_document(base_title, source_label, start_row, current[-1], current))
            current = []
            current_chars = 0
            current_count = 0
        if not current:
            start_row = row_number
        current.append(section)
        current_chars += section_chars
        current_count += 1
    if current:
        documents.append(make_row_document(base_title, source_label, start_row, current[-1], current))
    return documents


def make_row_document(
    base_title: str,
    source_label: str,
    start_row: int,
    last_section: str,
    sections: list[str],
) -> ImportedKnowledgeDocument:
    end_row = parse_section_row_number(last_section) or start_row
    title = f"{base_title} / {source_label} / rows {start_row}-{end_row}"[:255]
    content = "\n\n".join([f"Source: {source_label}", f"Rows: {start_row}-{end_row}", *sections])
    return ImportedKnowledgeDocument(title=title, content=content)


def parse_section_row_number(section: str) -> int | None:
    first_line = section.splitlines()[0] if section else ""
    if not first_line.startswith("Row "):
        return None
    value = first_line.removeprefix("Row ").strip()
    return int(value) if value.isdigit() else None


def format_structured_row(row_number: int, values: list[str], header: list[str]) -> str:
    lines = [f"Row {row_number}"]
    for index, value in enumerate(values):
        if not value:
            continue
        label = header[index].strip() if index < len(header) and header[index].strip() else f"Column {index + 1}"
        lines.append(f"{label}: {value}")
    return "\n".join(lines)


def split_text(content: str, limit: int) -> list[str]:
    lines = content.splitlines()
    sections: list[str] = []
    current: list[str] = []
    current_chars = 0
    for line in lines:
        line_chars = len(line) + 1
        if current and current_chars + line_chars > limit:
            sections.append("\n".join(current).strip())
            current = []
            current_chars = 0
        if line_chars > limit:
            sections.extend(line[index : index + limit] for index in range(0, len(line), limit))
            continue
        current.append(line)
        current_chars += line_chars
    if current:
        sections.append("\n".join(current).strip())
    return [section for section in sections if section]


def part_title(base_title: str, index: int, total: int) -> str:
    if total <= 1:
        return base_title[:255]
    return f"{base_title} / part {index}"[:255]


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def format_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_content(content: str) -> str:
    lines = [line.strip() for line in content.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    return "\n".join(line for line in lines if line).strip()
